import openpyxl
import abc
import csv
from typing import List, Dict, Any, Optional


class ExcelProcessFile(abc.ABC):
    """
    Clase base para procesar archivos Excel.
    """

    @abc.abstractmethod
    def read_data(self, file_path: str) -> List[Dict[str, Any]]:
        raise NotImplementedError

    @abc.abstractmethod
    def extract_relevant_data(self) -> List[Dict[str, Any]]:
        raise NotImplementedError

    @abc.abstractmethod
    def generate_csv(self) -> str:
        raise NotImplementedError
    
    @abc.abstractmethod
    def create_csv(self, top_10_data: List[Dict[str, Any]]) -> str:
        raise NotImplementedError


class ExcelAdapter(ExcelProcessFile):
    """
    Implementación de la clase ExcelProcessFile.
    """

    def __init__(self) -> None:
        self.data: Optional[List[Dict[str, Any]]] = None
        self.top_10: Optional[List[Dict[str, Any]]] = None

    def read_data(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Lee los datos del archivo Excel y los almacena en memoria.
        """
        wb_obj = openpyxl.load_workbook(file_path)
        sheet_obj = wb_obj[wb_obj.sheetnames[2]]

        self.data = [
            {
                "producto": sheet_obj.cell(row=row, column=5).value,
                "marca": sheet_obj.cell(row=row, column=7).value,
                "cantidad": sheet_obj.cell(row=row, column=8).value,
                "precio": sheet_obj.cell(row=row, column=11).value,
            }
            for row in range(9, 1355)
            if sheet_obj.cell(row=row, column=8).value is not None
        ]
        return self.data

    def extract_relevant_data(self) -> List[Dict[str, Any]]:
        """
        Extrae los 10 productos más vendidos.
        """
        if not self.data:
            raise ValueError("No hay datos cargados para extraer información.")
        self.top_10 = sorted(self.data, key=lambda x: x["cantidad"], reverse=True)[:10]
        
        return self.top_10

    def generate_csv(self) -> str:
        """
        Genera un archivo JSON y un CSV con los 10 productos más vendidos.
        """
        if self.top_10 is None:
            self.extract_relevant_data()

        __path__ = self.create_csv(self.top_10)
        
        return __path__

    def create_csv(self, top_10_data: List[Dict[str, Any]]) -> str:
        """
        Crea un archivo CSV basado en los datos de los 10 mejores productos
        y calcula el total de los precios.
        """

        csv_file = "downloads/top_10_best_sellers.csv"
        total_precio = 0  # Inicializamos la suma total de los precios

        with open(csv_file, mode="w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["Nombre del Producto", "Marca", "Cantidad", "Precio"])

            for item in top_10_data:
                try:
                    # Convertimos el precio a flotante
                    precio = float(item.get("precio", 0))
                except (ValueError, TypeError):
                    precio = 0

                total_precio += precio

                writer.writerow(
                    [
                        item.get("producto", "N/A"),
                        item.get("marca", "N/A"),
                        item.get("cantidad", 0),
                        precio,
                    ]
                )

            writer.writerow(["TOTAL", "", "", round(total_precio, 2)])
            
            return csv_file

